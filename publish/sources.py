"""
Source adapters.

Every kind of thing you can publish is described here and nowhere else. An
adapter knows how to (a) list what the user could publish, (b) fetch one, and
(c) freeze it into files at publish time.

Nothing at module level imports hanns / kura / chalk / cards. Models are
resolved lazily through django.apps and every adapter reports available() =
False if its app is not installed, in which case the UI simply does not offer
that kind. Point an adapter somewhere else with settings:

    PUBLISH_SOURCE_MODELS = {
        "deck":    "hanns.Deck",
        "dataset": "kura.Survey",
        "board":   "chalk.Board",
        "card":    "cards.Card",
        "show":    "polls.Session",
    }
"""

import csv
import io
import json

from django.apps import apps
from django.conf import settings
from django.utils import timezone

DEFAULT_SOURCE_MODELS = {
    "deck": "hanns.Deck",
    "dataset": "kura.Survey",
    "board": "chalk.Board",
    "card": "cards.Card",
    "show": "attendance.Event",
}

_REGISTRY = {}


def source_models():
    merged = dict(DEFAULT_SOURCE_MODELS)
    merged.update(getattr(settings, "PUBLISH_SOURCE_MODELS", {}) or {})
    return merged


def register(cls):
    _REGISTRY[cls.kind] = cls()
    return cls


def adapter_for(kind):
    return _REGISTRY.get(kind)


def available_adapters():
    return [a for a in _REGISTRY.values() if a.available()]


class SourceItem:
    """One row in the "pick what to publish" list."""

    def __init__(self, ref, title, subtitle="", updated=None, thumbnail=""):
        self.ref = str(ref)
        self.title = title or "Untitled"
        self.subtitle = subtitle
        self.updated = updated
        self.thumbnail = thumbnail

    def as_dict(self):
        return {
            "ref": self.ref,
            "title": self.title,
            "subtitle": self.subtitle,
            "updated": self.updated.isoformat() if self.updated else "",
            "thumbnail": self.thumbnail,
        }


class Asset:
    """A file to freeze. Give it bytes or text, never a path."""

    def __init__(self, role, label, filename, content, media_type="application/octet-stream",
                 order=0, row_count=None, column_count=None):
        if isinstance(content, str):
            content = content.encode("utf-8")
        self.role = role
        self.label = label
        self.filename = filename
        self.content = content
        self.media_type = media_type
        self.order = order
        self.row_count = row_count
        self.column_count = column_count


# --------------------------------------------------------------------------- #
# base
# --------------------------------------------------------------------------- #

class SourceAdapter:
    kind = ""
    label = ""
    noun = ""
    icon = "bi-file-earmark"
    blurb = ""
    needs_source = True
    variants = []          # e.g. [("clean", "Cleaned data"), ("raw", "Raw data")]

    # -- discovery ---------------------------------------------------------- #

    def model(self):
        path = source_models().get(self.kind)
        if not path:
            return None
        try:
            return apps.get_model(path)
        except Exception:
            return None

    def available(self):
        return not self.needs_source or self.model() is not None

    def owner_filter(self, user):
        """Best-effort 'mine' filter across the differing field names in KnockKnock."""
        from django.db.models import Q
        model = self.model()
        names = {f.name for f in model._meta.get_fields() if hasattr(f, "name")}
        q = Q()
        for field in ("owner", "user", "created_by", "author", "teacher", "host"):
            if field in names:
                q |= Q(**{field: user})
        return q

    def queryset_for(self, user):
        model = self.model()
        if model is None:
            return []
        try:
            qs = model.objects.filter(self.owner_filter(user))
        except Exception:
            qs = model.objects.none()
        for field in ("-updated_at", "-modified", "-created_at", "-id"):
            try:
                return qs.order_by(field)[:80]
            except Exception:
                continue
        return qs[:80]

    def list_for_user(self, user):
        out = []
        for obj in self.queryset_for(user):
            out.append(SourceItem(
                ref=obj.pk,
                title=self.title_of(obj),
                subtitle=self.subtitle_of(obj),
                updated=getattr(obj, "updated_at", None) or getattr(obj, "created_at", None),
            ))
        return out

    def fetch(self, user, ref):
        model = self.model()
        if model is None or not ref:
            return None
        try:
            obj = model.objects.get(pk=ref)
        except Exception:
            return None
        if user is not None and getattr(user, "is_staff", False):
            return obj
        for field in ("owner", "user", "created_by", "author", "teacher", "host"):
            holder = getattr(obj, field, None)
            if holder is not None:
                return obj if holder == user else None
        return obj

    # -- description --------------------------------------------------------- #

    def title_of(self, obj):
        for field in ("title", "name", "label"):
            val = getattr(obj, field, None)
            if val:
                return str(val)
        return str(obj)

    def subtitle_of(self, obj):
        return ""

    # -- freezing ------------------------------------------------------------ #

    def build_assets(self, publication, obj):
        """Return a list of Asset. Called once, at publish time."""
        return []

    def reader_context(self, publication):
        """Extra context the detail template can use to render the body."""
        return {}


# --------------------------------------------------------------------------- #
# article
# --------------------------------------------------------------------------- #

@register
class ArticleAdapter(SourceAdapter):
    kind = "article"
    label = "Article"
    noun = "article"
    icon = "bi-file-text"
    blurb = "Write it here. Headings, figures, tables, quotes and footnotes."
    needs_source = False

    def build_assets(self, publication, obj):
        return []


# --------------------------------------------------------------------------- #
# Kura dataset
# --------------------------------------------------------------------------- #

@register
class DatasetAdapter(SourceAdapter):
    kind = "dataset"
    label = "Dataset"
    noun = "Kura survey"
    icon = "bi-table"
    blurb = "Publish the responses from a Kura survey, with a codebook built from the form."
    variants = [("clean", "Cleaned data"), ("raw", "Raw data, exactly as collected")]

    def subtitle_of(self, obj):
        try:
            return "%d responses" % obj.submissions.count()
        except Exception:
            return ""

    # -- pulling rows out of Kura ------------------------------------------- #

    def _questions(self, survey):
        """
        Kura stores its questionnaire as a list of item dicts. Accept the shapes
        it is known to use and fall back to reading keys off the data.
        """
        for attr in ("questions", "schema", "definition", "form", "items"):
            val = getattr(survey, attr, None)
            if callable(val):
                try:
                    val = val()
                except Exception:
                    val = None
            if isinstance(val, dict):
                val = val.get("questions") or val.get("items")
            if isinstance(val, list) and val:
                return [q for q in val if isinstance(q, dict)]
        version = getattr(survey, "current_version", None) or getattr(survey, "latest_version", None)
        if version is not None:
            for attr in ("questions", "schema", "definition"):
                val = getattr(version, attr, None)
                if isinstance(val, list) and val:
                    return [q for q in val if isinstance(q, dict)]
        return []

    def _submissions(self, survey, variant):
        try:
            qs = survey.submissions.all()
        except Exception:
            return []
        try:
            if variant == "clean":
                for field in ("is_clean", "cleaned"):
                    if any(f.name == field for f in qs.model._meta.get_fields()):
                        qs = qs.filter(**{field: True})
                        break
                else:
                    # No boolean, so fall back to dropping flagged rows.
                    try:
                        qs = qs.exclude(flags__resolved=False)
                    except Exception:
                        pass
        except Exception:
            pass
        try:
            return list(qs.order_by("id")[: getattr(settings, "PUBLISH_MAX_ROWS", 200000)])
        except Exception:
            return list(qs[:200000])

    def _row_data(self, sub, variant):
        for attr in (("clean_data", "data") if variant == "clean" else ("data", "answers", "payload")):
            val = getattr(sub, attr, None)
            if isinstance(val, dict) and val:
                return val
        val = getattr(sub, "answers", None)
        return val if isinstance(val, dict) else {}

    def frame(self, survey, variant):
        """-> (columns, rows) with rows as lists of strings."""
        questions = self._questions(survey)
        subs = self._submissions(survey, variant)
        cols = []
        for q in questions:
            name = q.get("name") or q.get("id") or q.get("code")
            if name and name not in cols and q.get("type") not in ("note", "section"):
                cols.append(name)
        seen = set(cols)
        for sub in subs:
            for key in self._row_data(sub, variant).keys():
                if key not in seen:
                    seen.add(key)
                    cols.append(key)
        meta_cols = ["_id", "_submitted_at"]
        rows = []
        for sub in subs:
            data = self._row_data(sub, variant)
            row = [str(sub.pk), self._when(sub)]
            for c in cols:
                row.append(self._flatten(data.get(c)))
            rows.append(row)
        return meta_cols + cols, rows

    @staticmethod
    def _when(sub):
        for attr in ("received_at", "submitted_at", "created_at"):
            val = getattr(sub, attr, None)
            if val:
                try:
                    return timezone.localtime(val).isoformat(timespec="seconds")
                except Exception:
                    return str(val)
        return ""

    @staticmethod
    def _flatten(value):
        if value is None:
            return ""
        if isinstance(value, (list, tuple)):
            if value and isinstance(value[0], dict):
                return json.dumps(value, ensure_ascii=False)
            return " ".join(str(v) for v in value)
        if isinstance(value, dict):
            return json.dumps(value, ensure_ascii=False)
        return str(value)

    def codebook(self, survey):
        entries = []
        for q in self._questions(survey):
            name = q.get("name") or q.get("id") or q.get("code")
            if not name:
                continue
            choices = q.get("choices") or q.get("options") or []
            entries.append({
                "variable": name,
                "label": q.get("label") or q.get("title") or q.get("text") or "",
                "type": q.get("type") or "text",
                "required": bool(q.get("required")),
                "asked_when": q.get("show") or q.get("relevant") or "",
                "calculation": q.get("calculate") or q.get("calculation") or "",
                "repeat_group": q.get("repeat") or q.get("parent") or "",
                "choices": [
                    {"value": c.get("value", c.get("name", c)) if isinstance(c, dict) else c,
                     "label": c.get("label", "") if isinstance(c, dict) else ""}
                    for c in choices
                ],
            })
        return entries

    def build_assets(self, publication, survey):
        variant = publication.source_variant or "clean"
        cols, rows = self.frame(survey, variant)

        buf = io.StringIO()
        writer = csv.writer(buf, lineterminator="\n")
        writer.writerow(cols)
        writer.writerows(rows)
        csv_bytes = buf.getvalue()

        book = self.codebook(survey)
        assets = [
            Asset("primary", "Data (CSV)", "%s-%s.csv" % (publication.slug[:60], variant),
                  csv_bytes, "text/csv", order=0, row_count=len(rows), column_count=len(cols)),
            Asset("codebook", "Codebook (JSON)", "%s-codebook.json" % publication.slug[:60],
                  json.dumps({"variables": book}, indent=2, ensure_ascii=False),
                  "application/json", order=2),
            Asset("codebook", "Codebook (CSV)", "%s-codebook.csv" % publication.slug[:60],
                  self._codebook_csv(book), "text/csv", order=3),
        ]
        xlsx = self._xlsx(cols, rows, book)
        if xlsx:
            assets.append(Asset(
                "data", "Data (Excel)", "%s-%s.xlsx" % (publication.slug[:60], variant), xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", order=1,
                row_count=len(rows), column_count=len(cols),
            ))
        publication.meta["rows"] = len(rows)
        publication.meta["columns"] = len(cols)
        publication.meta["variables"] = len(book)
        publication.meta["preview"] = {"columns": cols[:12], "rows": [r[:12] for r in rows[:25]]}
        return assets

    @staticmethod
    def _codebook_csv(book):
        buf = io.StringIO()
        w = csv.writer(buf, lineterminator="\n")
        w.writerow(["variable", "label", "type", "required", "asked_when", "choices"])
        for e in book:
            choices = "; ".join(
                "%s=%s" % (c["value"], c["label"]) if c["label"] else str(c["value"])
                for c in e["choices"]
            )
            w.writerow([e["variable"], e["label"], e["type"], "yes" if e["required"] else "",
                        e["asked_when"], choices])
        return buf.getvalue()

    @staticmethod
    def _xlsx(cols, rows, book):
        try:
            from openpyxl import Workbook
        except Exception:
            return None
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("data")
        ws.append(cols)
        for r in rows:
            ws.append(r)
        cb = wb.create_sheet("codebook")
        cb.append(["variable", "label", "type", "required", "asked_when"])
        for e in book:
            cb.append([e["variable"], e["label"], e["type"], "yes" if e["required"] else "", e["asked_when"]])
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    def reader_context(self, publication):
        return {"preview": (publication.meta or {}).get("preview") or {}}


# --------------------------------------------------------------------------- #
# Hanns deck
# --------------------------------------------------------------------------- #

@register
class DeckAdapter(SourceAdapter):
    kind = "deck"
    label = "Slide deck"
    noun = "Hanns deck"
    icon = "bi-easel"
    blurb = "Publish a deck as a player readers can step through, plus a PowerPoint download."

    def subtitle_of(self, obj):
        try:
            return "%d slides" % obj.slides.count()
        except Exception:
            return ""

    def _call_exporter(self, setting_name, default_paths, deck):
        from django.utils.module_loading import import_string
        paths = [getattr(settings, setting_name, None)] + list(default_paths)
        for path in paths:
            if not path:
                continue
            try:
                fn = import_string(path)
            except Exception:
                continue
            try:
                return fn(deck)
            except Exception:
                continue
        return None

    def build_assets(self, publication, deck):
        assets = []
        html = self._call_exporter(
            "PUBLISH_HANNS_HTML_EXPORT",
            ["hanns.html_exporter.export_deck", "hanns.html_exporter.build_html",
             "hanns.html_exporter.export"],
            deck,
        )
        if html:
            assets.append(Asset("primary", "Deck player (HTML)", "%s.html" % publication.slug[:60],
                                html, "text/html", order=0))
        pptx = self._call_exporter(
            "PUBLISH_HANNS_PPTX_EXPORT",
            ["hanns.pptx_exporter.export_deck", "hanns.powerpoint_exporter.export_deck",
             "hanns.exporters.export_pptx"],
            deck,
        )
        if pptx:
            assets.append(Asset(
                "extra", "PowerPoint (PPTX)", "%s.pptx" % publication.slug[:60], pptx,
                "application/vnd.openxmlformats-officedocument.presentationml.presentation", order=1))
        try:
            publication.meta["slides"] = deck.slides.count()
        except Exception:
            pass
        return assets

    def reader_context(self, publication):
        asset = publication.assets.filter(version=publication.version, role="primary").first()
        return {"player_asset": asset}


# --------------------------------------------------------------------------- #
# Chalk board
# --------------------------------------------------------------------------- #

@register
class BoardAdapter(SourceAdapter):
    kind = "board"
    label = "Board"
    noun = "Chalk board"
    icon = "bi-easel2"
    blurb = "Publish a lesson board. Every page is frozen as vector artwork, not a screenshot."

    def subtitle_of(self, obj):
        try:
            return "%d pages" % obj.pages.count()
        except Exception:
            return ""

    SURFACE_BG = {
        "black": "#12130f", "green": "#123324", "white": "#ffffff",
        "grid": "#fbfbf8", "ruled": "#fdfcf6",
    }

    def build_assets(self, publication, board):
        try:
            pages = list(board.pages.all().order_by("number"))
        except Exception:
            pages = []
        assets = []
        svgs = []
        for i, page in enumerate(pages, start=1):
            svg = self.page_svg(board, page)
            svgs.append(svg)
            assets.append(Asset("extra" if i > 1 else "primary",
                                "Page %d (SVG)" % i,
                                "%s-p%d.svg" % (publication.slug[:50], i),
                                svg, "image/svg+xml", order=i))
        if svgs:
            publication.meta["pages"] = len(svgs)
            publication.meta["board_svg"] = svgs[0][: 400000]
        return assets

    def page_svg(self, board, page, width=1600, height=900):
        surface = getattr(page, "surface", None) or getattr(board, "surface", "black")
        bg = self.SURFACE_BG.get(str(surface), "#12130f")
        parts = [
            '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 %d %d" width="%d" height="%d" '
            'role="img">' % (width, height, width, height),
            '<rect width="100%%" height="100%%" fill="%s"/>' % bg,
        ]
        for stroke in self._strokes(page):
            path = self._stroke_path(stroke, width, height)
            if path:
                parts.append(path)
        for el in self._els(page):
            frag = self._el_svg(el, width, height)
            if frag:
                parts.append(frag)
        parts.append("</svg>")
        return "".join(parts)

    @staticmethod
    def _strokes(page):
        val = getattr(page, "strokes", None)
        if isinstance(val, list):
            return [s for s in val if isinstance(s, dict)]
        if isinstance(val, dict):
            return [s for s in val.get("strokes", []) if isinstance(s, dict)]
        return []

    @staticmethod
    def _els(page):
        val = getattr(page, "els", None)
        return [e for e in val if isinstance(e, dict)] if isinstance(val, list) else []

    @staticmethod
    def _stroke_path(stroke, W, H):
        pts = stroke.get("p") or stroke.get("pts") or stroke.get("points") or []
        flat = []
        if pts and isinstance(pts[0], (int, float)):
            flat = list(pts)
        else:
            for p in pts:
                if isinstance(p, dict):
                    flat += [p.get("x", 0), p.get("y", 0)]
                elif isinstance(p, (list, tuple)) and len(p) >= 2:
                    flat += [p[0], p[1]]
        if len(flat) < 4:
            return ""
        d = []
        for i in range(0, len(flat) - 1, 2):
            x, y = float(flat[i]) * W, float(flat[i + 1]) * H
            d.append(("M" if i == 0 else "L") + "%.1f %.1f" % (x, y))
        colour = stroke.get("c") or stroke.get("color") or "#f3f2ea"
        width = float(stroke.get("w") or stroke.get("width") or 0.004) * W
        opacity = stroke.get("o") or stroke.get("opacity") or 1
        return ('<path d="%s" fill="none" stroke="%s" stroke-width="%.2f" stroke-opacity="%s" '
                'stroke-linecap="round" stroke-linejoin="round"/>'
                % ("".join(d), _xml(colour), max(0.5, width), opacity))

    @staticmethod
    def _el_svg(el, W, H):
        from django.utils.html import escape
        x, y = float(el.get("x", 0)) * W, float(el.get("y", 0)) * H
        w, h = float(el.get("w", 0.2)) * W, float(el.get("h", 0.1)) * H
        etype = el.get("type", "")
        if etype == "text":
            size = float(el.get("size", 0.04)) * H
            return ('<text x="%.1f" y="%.1f" fill="%s" font-size="%.1f" '
                    'font-family="Georgia, serif">%s</text>'
                    % (x, y + size, _xml(el.get("color", "#f3f2ea")), size, escape(el.get("text", ""))))
        if etype == "image" and el.get("src"):
            return '<image x="%.1f" y="%.1f" width="%.1f" height="%.1f" href="%s"/>' % (
                x, y, w, h, _xml(el["src"]))
        stroke = _xml(el.get("stroke", el.get("color", "#f3f2ea")))
        fill = el.get("fill") or "none"
        return ('<rect x="%.1f" y="%.1f" width="%.1f" height="%.1f" fill="%s" stroke="%s" '
                'stroke-width="2" rx="6"/>' % (x, y, w, h, _xml(fill), stroke))

    def reader_context(self, publication):
        return {"board_svg": (publication.meta or {}).get("board_svg", "")}


def _xml(value):
    return (str(value).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


# --------------------------------------------------------------------------- #
# Card
# --------------------------------------------------------------------------- #

@register
class CardAdapter(SourceAdapter):
    kind = "card"
    label = "Card"
    noun = "group card"
    icon = "bi-card-heading"
    blurb = "Publish a group card with its wall of messages."

    def build_assets(self, publication, card):
        messages = []
        for rel in ("messages", "notes", "entries", "signatures"):
            qs = getattr(card, rel, None)
            if qs is None:
                continue
            try:
                for m in qs.all()[:500]:
                    messages.append({
                        "from": str(getattr(m, "author_name", "") or getattr(m, "name", "") or ""),
                        "text": str(getattr(m, "text", "") or getattr(m, "body", "") or ""),
                        "at": str(getattr(m, "created_at", "") or ""),
                    })
                break
            except Exception:
                continue
        publication.meta["messages"] = messages[:200]
        publication.meta["message_count"] = len(messages)
        return [Asset("primary", "Card contents (JSON)", "%s.json" % publication.slug[:60],
                      json.dumps({"title": self.title_of(card), "messages": messages},
                                 indent=2, ensure_ascii=False),
                      "application/json", order=0)]

    def reader_context(self, publication):
        return {"card_messages": (publication.meta or {}).get("messages", [])}


# --------------------------------------------------------------------------- #
# Show
# --------------------------------------------------------------------------- #

@register
class ShowAdapter(SourceAdapter):
    kind = "show"
    label = "Show"
    noun = "live session"
    icon = "bi-broadcast"
    blurb = "Publish the record of a live session: what was run, who came, what the room answered."

    def subtitle_of(self, obj):
        for field in ("starts_at", "date", "created_at"):
            val = getattr(obj, field, None)
            if val:
                return str(val)[:16]
        return ""

    def build_assets(self, publication, obj):
        record = {"title": self.title_of(obj), "captured": timezone.now().isoformat()}
        for field in ("starts_at", "ends_at", "venue", "location", "description", "summary"):
            val = getattr(obj, field, None)
            if val:
                record[field] = str(val)
        for rel, key in (("sessions", "sessions"), ("polls", "polls"),
                         ("questions", "questions"), ("attendees", "attendees")):
            qs = getattr(obj, rel, None)
            if qs is None:
                continue
            try:
                record[key] = [str(x) for x in qs.all()[:500]]
                if key == "attendees":
                    record["attendee_count"] = qs.count()
            except Exception:
                continue
        publication.meta["show"] = {k: v for k, v in record.items() if not isinstance(v, list)}
        publication.meta["show_lists"] = {k: v for k, v in record.items() if isinstance(v, list)}
        return [Asset("primary", "Session record (JSON)", "%s.json" % publication.slug[:60],
                      json.dumps(record, indent=2, ensure_ascii=False), "application/json", order=0)]

    def reader_context(self, publication):
        meta = publication.meta or {}
        return {"show": meta.get("show", {}), "show_lists": meta.get("show_lists", {})}
